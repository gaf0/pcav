import streamlit as st
import pandas as pd
import random
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- CONFIGURATION ---
COLORS = {
    "Sous - sol": "DDEBF7", 
    "RDC": "FCE4D6",      
    "1st floor": "E2EFDA", 
    "Header": "F2F2F2"    
}

st.set_page_config(page_title="Lab Schedule Generator", layout="wide")
st.title("🧪 Lab Task Draw Generator")

uploaded_file = st.file_uploader("Choose your Excel input file", type="xlsx")

if uploaded_file:
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", datetime.now())
    with col2:
        end_date = st.date_input("End Date", datetime.now() + timedelta(weeks=12))

    if st.button("Generate & Style Schedule"):
        try:
            # 1. Load Data
            df = pd.read_excel(uploaded_file, header=[0, 1])
            
            # Reconstruct Headers
            raw_cols = df.columns.to_frame()
            raw_cols[0] = raw_cols[0].mask(raw_cols[0].str.contains('Unnamed')).ffill()
            df.columns = [f"{col[0]} - {col[1]}" if "Unnamed" not in str(col[1]) else col[0] for col in raw_cols.values]
            
            name_col, contract_col, task_cols = df.columns[0], df.columns[2], df.columns[3:]

            # FIX: Robust Date Parsing for "Mar.26" format
            def parse_contract(x):
                try:
                    # Assumes format "Mon.YY" like Mar.26
                    return datetime.strptime(str(x), "%b.%y").date()
                except:
                    return datetime.max.date()
            
            df['expiry'] = df[contract_col].apply(parse_contract)

            # 2. Draw Logic
            buckets = {task: [] for task in task_cols}
            schedule_data = []
            curr = start_date - timedelta(days=start_date.weekday()) # Monday anchor

            while curr <= end_date:
                week_label = f"W{curr.isocalendar()[1]} ({curr.strftime('%d/%m/%Y')})"
                week_row = {"Week": week_label}
                picked_this_week = set()

                for task in task_cols:
                    # Eligibility Check: Active contract AND Ticked box
                    eligible = df[
                        (df[task].notna()) & 
                        (df[task] != False) & 
                        (df['expiry'] >= curr) # MUST be active during this specific week
                    ][name_col].tolist()
                    
                    if not eligible:
                        week_row[task] = "N/A (No active contracts)"
                        continue
                    
                    # Refill bucket logic
                    valid_in_bucket = [p for p in buckets[task] if p in eligible]
                    if not valid_in_bucket:
                        buckets[task] = list(eligible)
                        random.shuffle(buckets[task])
                        valid_in_bucket = buckets[task]

                    # Pick winner prioritizing those not already picked this week
                    fresh = [p for p in valid_in_bucket if p not in picked_this_week]
                    winner = random.choice(fresh if fresh else valid_in_bucket)
                    
                    buckets[task].remove(winner)
                    picked_this_week.add(winner)
                    week_row[task] = winner
                    
                schedule_data.append(week_row)
                curr += timedelta(days=7)

            # 3. Create Excel
            output = BytesIO()
            final_df = pd.DataFrame(schedule_data)
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Schedule')
                ws = writer.sheets['Schedule']
                
                # Apply same styling logic
                for col_num, col_title in enumerate(final_df.columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    
                    fill_color = COLORS["Header"]
                    if "Sous - sol" in col_title: fill_color = COLORS["Sous - sol"]
                    elif "RDC" in col_title: fill_color = COLORS["RDC"]
                    elif "1st floor" in col_title: fill_color = COLORS["1st floor"]
                    
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    ws.column_dimensions[cell.column_letter].width = 25

            st.success("Corrected Schedule Generated!")
            st.dataframe(final_df) # Visual preview to verify Yasmine's absence after March
            
            st.download_button(
                label="📥 Download Corrected Schedule",
                data=output.getvalue(),
                file_name=f"Corrected_Lab_Schedule_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error: {e}")